#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
암호화폐 매수 신호 데이터 분석 시스템 (5분봉 급등 감지)
- 초단타 급등 패턴 수집 및 분석
- 시장 데이터 히스토리 데이터베이스 구축
- 분석 리포트 자동 생성 및 Git 저장
"""

import pyupbit
import pandas as pd
import numpy as np
import requests
import time
from datetime import datetime, timedelta
import pytz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
import os
import json
warnings.filterwarnings('ignore')

KST = pytz.timezone('Asia/Seoul')

def get_kst_now():
    return datetime.now(KST)

# ============================================
# 환경변수 설정
# ============================================
BOT_TOKEN = os.environ.get('BOT_TOKEN', '')
CHAT_ID = os.environ.get('CHAT_ID', '')

SCAN_INTERVAL = int(os.environ.get('SCAN_INTERVAL', '120'))
VOLUME_SPIKE_THRESHOLD = float(os.environ.get('VOLUME_SPIKE_THRESHOLD', '1.8'))
PRICE_CHANGE_THRESHOLD = float(os.environ.get('PRICE_CHANGE_THRESHOLD', '2.5'))
CONSECUTIVE_THRESHOLD = int(os.environ.get('CONSECUTIVE_THRESHOLD', '2'))

# 데이터 저장 경로 (Repository 내)
DATA_DIR = 'market_data/buy_signals'
ANALYSIS_DIR = 'analysis_reports/buy_reports'
EXCEL_FILE = 'buy_signals_database.xlsx'

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(ANALYSIS_DIR, exist_ok=True)

# ============================================
# 데이터 수집 및 분석
# ============================================

def collect_market_data():
    """시장 데이터 수집"""
    print(f"📊 급등 신호 데이터 수집: {get_kst_now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    tickers = pyupbit.get_tickers(fiat="KRW")
    market_snapshot = []
    
    for coin in tickers:
        try:
            analysis = detect_price_surge(coin)
            if analysis:
                market_snapshot.append(analysis)
            time.sleep(0.05)
        except Exception as e:
            continue
    
    return market_snapshot

def detect_price_surge(coin):
    """5분봉 기반 급등 조기 감지"""
    try:
        df = pyupbit.get_ohlcv(coin, interval="minute5", count=50)
        if df is None or len(df) < 20:
            return None
        
        current_candle = df.iloc[-1]
        current_volume = current_candle['volume']
        current_price = current_candle['close']
        
        # 거래량 분석
        avg_volume = df['volume'].iloc[-11:-1].mean()
        volume_ratio = current_volume / avg_volume if avg_volume > 0 else 0
        
        recent_3_volume = df['volume'].iloc[-3:].sum()
        prev_10_volume = df['volume'].iloc[-13:-3].sum()
        volume_acceleration = recent_3_volume / prev_10_volume if prev_10_volume > 0 else 0
        
        # 가격 분석
        candle_change = ((current_candle['close'] - current_candle['open']) / current_candle['open']) * 100
        
        price_5m_ago = df['close'].iloc[-2]
        price_change_5m = ((current_price - price_5m_ago) / price_5m_ago) * 100
        
        if len(df) >= 4:
            price_15m_ago = df['close'].iloc[-4]
            price_change_15m = ((current_price - price_15m_ago) / price_15m_ago) * 100
        else:
            price_change_15m = 0
        
        # 연속 상승 분석
        consecutive_green = 0
        for i in range(1, min(6, len(df))):
            if df['close'].iloc[-i] > df['open'].iloc[-i]:
                consecutive_green += 1
            else:
                break
        
        consecutive_volume = 0
        for i in range(1, min(5, len(df))):
            if df['volume'].iloc[-i] > df['volume'].iloc[-i-1]:
                consecutive_volume += 1
            else:
                break
        
        # 체결강도
        recent_5 = df.iloc[-5:]
        green_count = sum(recent_5['close'] > recent_5['open'])
        buying_pressure = green_count / 5
        
        high_20 = df['high'].iloc[-21:-1].max()
        breaking_high = current_price > high_20
        
        # 호가창
        orderbook_data = analyze_orderbook_momentum(coin)
        
        return {
            'timestamp': get_kst_now().isoformat(),
            'coin': coin,
            'price': float(current_price),
            'volume': float(current_volume),
            'volume_ratio': float(volume_ratio),
            'volume_acceleration': float(volume_acceleration),
            'candle_change': float(candle_change),
            'price_change_5m': float(price_change_5m),
            'price_change_15m': float(price_change_15m),
            'consecutive_green': int(consecutive_green),
            'consecutive_volume': int(consecutive_volume),
            'buying_pressure': float(buying_pressure),
            'breaking_high': bool(breaking_high),
            'orderbook': orderbook_data
        }
    except Exception as e:
        return None

def analyze_orderbook_momentum(coin):
    """호가창 매수/매도 압력 분석"""
    try:
        orderbook = pyupbit.get_orderbook(coin)
        if not orderbook or not isinstance(orderbook, list):
            return None
        
        ob = orderbook[0]
        if 'orderbook_units' not in ob:
            return None
        
        units = ob['orderbook_units']
        
        total_bid = sum([u.get('bid_size', 0) for u in units])
        total_ask = sum([u.get('ask_size', 0) for u in units])
        
        top3_bid = sum([units[i].get('bid_size', 0) for i in range(min(3, len(units)))])
        top3_ask = sum([units[i].get('ask_size', 0) for i in range(min(3, len(units)))])
        
        bid_ask_ratio = total_bid / total_ask if total_ask > 0 else 0
        top3_ratio = top3_bid / top3_ask if top3_ask > 0 else 0
        
        imbalance = (total_bid - total_ask) / (total_bid + total_ask) if (total_bid + total_ask) > 0 else 0
        
        return {
            'bid_ask_ratio': float(bid_ask_ratio),
            'top3_ratio': float(top3_ratio),
            'imbalance': float(imbalance),
            'total_bid': float(total_bid),
            'total_ask': float(total_ask)
        }
    except Exception as e:
        return None

# ============================================
# 신호 평가
# ============================================

def evaluate_fast_signal(surge_data):
    """초단타 신호 강도 평가 (0-10점)"""
    score = 0
    signals = []
    alert_level = "NORMAL"
    
    if not surge_data:
        return 0, [], "NONE"
    
    # 거래량 폭발 (0-3점)
    if surge_data['volume_ratio'] >= 3.0:
        score += 3
        signals.append("🔥🔥 거래량 3배 폭발")
        alert_level = "CRITICAL"
    elif surge_data['volume_ratio'] >= 2.0:
        score += 2
        signals.append("🔥 거래량 2배 급증")
        alert_level = "HIGH"
    elif surge_data['volume_ratio'] >= 1.5:
        score += 1
        signals.append("⚡ 거래량 1.5배 증가")
    
    # 가격 급등 (0-3점)
    if surge_data['price_change_5m'] >= 5:
        score += 3
        signals.append("🚀🚀 5분 5% 급등")
        alert_level = "CRITICAL"
    elif surge_data['price_change_5m'] >= 3:
        score += 2
        signals.append("🚀 5분 3% 상승")
        if alert_level == "NORMAL":
            alert_level = "HIGH"
    elif surge_data['price_change_5m'] >= 2:
        score += 1
        signals.append("📈 5분 2% 상승")
    
    # 연속 상승 (0-2점)
    if surge_data['consecutive_green'] >= 4:
        score += 2
        signals.append("✅ 4연속 양봉")
    elif surge_data['consecutive_green'] >= 3:
        score += 1
        signals.append("✅ 3연속 양봉")
    
    # 거래량 가속 (0-1점)
    if surge_data['volume_acceleration'] >= 2.0:
        score += 1
        signals.append("⚡ 거래량 가속")
    
    # 매수세 우위 (0-1점)
    if surge_data['buying_pressure'] >= 0.8:
        score += 1
        signals.append("💪 강한 매수세")
    
    # 고점 돌파 (0-1점)
    if surge_data['breaking_high']:
        score += 1
        signals.append("🎯 20봉 고점 돌파")
    
    # 호가창 매수세 (0-1점)
    orderbook = surge_data.get('orderbook', {})
    if orderbook and orderbook.get('bid_ask_ratio', 0) >= 1.8:
        score += 1
        signals.append("💰 호가창 매수벽")
    
    return score, signals, alert_level

# ============================================
# 데이터 저장 (Repository 활용)
# ============================================

def save_to_json_history(market_snapshot):
    """JSON 히스토리 저장"""
    history_file = os.path.join(DATA_DIR, 'buy_signals_history.json')
    
    try:
        if os.path.exists(history_file):
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
        else:
            history = []
        
        history.append({
            'scan_time': get_kst_now().isoformat(),
            'data': market_snapshot
        })
        
        if len(history) > 100:
            history = history[-100:]
        
        with open(history_file, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        
        print(f"✅ JSON 저장: {len(market_snapshot)}개")
        return True
    except Exception as e:
        print(f"❌ JSON 저장 실패: {e}")
        return False

def save_to_excel_database(market_snapshot):
    """Excel 데이터베이스 저장"""
    try:
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = "급등신호"
            
            headers = ['수집시간', '코인', '레벨', '점수', '현재가', '거래량배수', 
                      '5분변화%', '15분변화%', '연속양봉', '매수세%']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        scan_time = get_kst_now().strftime('%Y-%m-%d %H:%M')
        for item in market_snapshot:
            score, signals, alert_level = evaluate_fast_signal(item)
            
            row = [
                scan_time,
                item['coin'].replace('KRW-', ''),
                alert_level,
                f"{score}/10",
                item['price'],
                f"{item['volume_ratio']:.2f}",
                f"{item['price_change_5m']:+.2f}",
                f"{item['price_change_15m']:+.2f}",
                item['consecutive_green'],
                f"{item['buying_pressure']*100:.0f}"
            ]
            ws.append(row)
        
        if ws.max_row > 1001:
            ws.delete_rows(2, ws.max_row - 1001)
        
        wb.save(EXCEL_FILE)
        print(f"✅ Excel 저장 완료")
        return True
    except Exception as e:
        print(f"❌ Excel 저장 실패: {e}")
        return False

def generate_buy_signal_report(market_snapshot):
    """매수 신호 리포트 생성"""
    report_date = get_kst_now().strftime('%Y%m%d_%H%M')
    report_path = os.path.join(ANALYSIS_DIR, f'buy_report_{report_date}.md')
    
    signals = []
    for item in market_snapshot:
        score, sig_list, alert_level = evaluate_fast_signal(item)
        if score >= 6:
            signals.append((item, score, alert_level))
    
    signals.sort(key=lambda x: x[1], reverse=True)
    
    report = f"""# 급등 매수 신호 분석 리포트

생성시간: {get_kst_now().strftime('%Y-%m-%d %H:%M:%S')}

## 📊 스캔 요약

- 분석 코인 수: {len(market_snapshot)}개
- 급등 신호 감지: {len(signals)}개

## 🎯 주요 급등 신호

"""
    
    for item, score, alert_level in signals[:20]:
        coin_name = item['coin'].replace('KRW-', '')
        report += f"""### {coin_name} (신호강도: {score}/10, {alert_level})

- 현재가: {item['price']:,.0f}원
- 거래량 배수: {item['volume_ratio']:.2f}배
- 5분 변화: {item['price_change_5m']:+.2f}%
- 15분 변화: {item['price_change_15m']:+.2f}%
- 연속 양봉: {item['consecutive_green']}개
- 매수세: {item['buying_pressure']*100:.0f}%

"""
    
    report += f"""
## 📈 시장 통계

- 평균 거래량 배수: {np.mean([x['volume_ratio'] for x in market_snapshot]):.2f}
- 평균 5분 변화율: {np.mean([x['price_change_5m'] for x in market_snapshot]):+.2f}%
- 고점 돌파 코인 수: {sum([1 for x in market_snapshot if x['breaking_high']])}개

---
*본 리포트는 자동 생성되었습니다.*
"""
    
    try:
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report)
        print(f"✅ 리포트 생성: {report_path}")
        return report_path, len(signals)
    except Exception as e:
        print(f"❌ 리포트 생성 실패: {e}")
        return None, 0

# ============================================
# Telegram 알림 (부가 기능)
# ============================================

def send_summary_notification(signals_count, report_path):
    """요약 알림"""
    if not BOT_TOKEN or not CHAT_ID:
        return
    
    try:
        message = f"""📊 급등 신호 분석 완료

⏰ {get_kst_now().strftime('%Y-%m-%d %H:%M:%S')}
🎯 급등 신호: {signals_count}개
📈 리포트: {os.path.basename(report_path) if report_path else 'N/A'}

데이터는 Repository에 저장되었습니다."""

        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        data = {"chat_id": CHAT_ID, "text": message}
        requests.post(url, data=data, timeout=10)
    except Exception as e:
        print(f"알림 전송 실패: {e}")

# ============================================
# Git Commit & Push
# ============================================

def commit_and_push_data():
    """Git 커밋 및 푸시"""
    try:
        import subprocess
        
        subprocess.run(['git', 'config', '--global', 'user.email', 'github-actions[bot]@users.noreply.github.com'], check=True)
        subprocess.run(['git', 'config', '--global', 'user.name', 'GitHub Actions Bot'], check=True)
        
        subprocess.run(['git', 'add', DATA_DIR, ANALYSIS_DIR, EXCEL_FILE], check=True)
        
        commit_msg = f"Update buy signals - {get_kst_now().strftime('%Y-%m-%d %H:%M')}"
        result = subprocess.run(['git', 'commit', '-m', commit_msg], capture_output=True, text=True)
        
        if result.returncode == 0:
            subprocess.run(['git', 'push'], check=True)
            print("✅ Git 커밋 및 푸시 완료")
            return True
        else:
            print("ℹ️ 변경사항 없음")
            return False
            
    except Exception as e:
        print(f"⚠️ Git 작업 실패: {e}")
        return False

# ============================================
# 메인 실행
# ============================================

def main():
    """메인"""
    print("""
    ╔══════════════════════════════════════╗
    ║   급등 신호 데이터 분석 시스템      ║
    ║   Fast Surge Signal Analysis        ║
    ╚══════════════════════════════════════╝
    """)
    
    try:
        # 1. 데이터 수집
        market_snapshot = collect_market_data()
        
        if not market_snapshot:
            print("❌ 수집된 데이터 없음")
            return
        
        # 2. 데이터 저장
        save_to_json_history(market_snapshot)
        save_to_excel_database(market_snapshot)
        
        # 3. 리포트 생성
        report_path, signals_count = generate_buy_signal_report(market_snapshot)
        
        # 4. Git 커밋
        commit_and_push_data()
        
        # 5. 알림
        send_summary_notification(signals_count, report_path)
        
        print(f"\n✅ 분석 완료: {len(market_snapshot)}개 코인, {signals_count}개 신호")
        
    except KeyboardInterrupt:
        print("\n🛑 프로그램 종료")
    except Exception as e:
        print(f"❌ 오류 발생: {e}")

if __name__ == "__main__":
    main()
