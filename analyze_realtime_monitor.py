#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
암호화폐 실시간 모니터링 데이터 분석 시스템
- 단기(5/15분봉) + 일봉 병행 분석
- 다중 기술적 지표 통합 분석
- 분석 리포트 자동 생성 및 Git 저장
"""

import pyupbit
import pandas as pd
import numpy as np
import requests
import time
from datetime import datetime, timedelta
import pytz
import ta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
import os
import json
warnings.filterwarnings('ignore')

KST = pytz.timezone('Asia/Seoul')

def get_kst_now():
    return datetime.now(KST)

# ============================================
# 환경변수 설정
# ============================================
BOT_TOKEN = os.environ.get('BOT_TOKEN', '')
CHAT_ID = os.environ.get('CHAT_ID', '')

SCAN_INTERVAL = int(os.environ.get('SCAN_INTERVAL', '180'))
VOLUME_THRESHOLD_WATCH = float(os.environ.get('VOLUME_THRESHOLD_WATCH', '1.3'))
VOLUME_THRESHOLD_STRONG = float(os.environ.get('VOLUME_THRESHOLD_STRONG', '2.0'))

# 데이터 저장 경로
DATA_DIR = 'market_data/realtime_monitor'
ANALYSIS_DIR = 'analysis_reports/realtime_reports'
EXCEL_FILE = 'realtime_monitor_database.xlsx'

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(ANALYSIS_DIR, exist_ok=True)

# ============================================
# 데이터 수집
# ============================================

def collect_market_data():
    """시장 데이터 수집"""
    print(f"📊 실시간 모니터링 데이터 수집: {get_kst_now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    tickers = pyupbit.get_tickers(fiat="KRW")
    market_snapshot = []
    
    for coin in tickers:
        try:
            analysis = analyze_coin_comprehensive(coin)
            if analysis:
                market_snapshot.append(analysis)
            time.sleep(0.1)
        except Exception as e:
            continue
    
    return market_snapshot

def analyze_coin_comprehensive(coin):
    """코인 종합 분석 (단기+일봉+지표)"""
    try:
        # 단기 시간봉 분석
        short_term_data = analyze_short_term_volume(coin)
        
        # 일봉 분석
        volume_data = analyze_volume(coin)
        
        # 호가창 분석
        orderbook_data = analyze_orderbook(coin)
        
        # 기술적 지표
        indicators = calculate_indicators(coin)
        
        # 신호 강도 계산
        score, signals, signal_type = calculate_signal_strength(
            volume_data, indicators, orderbook_data, short_term_data
        )
        
        if not short_term_data or not volume_data:
            return None
        
        return {
            'timestamp': get_kst_now().isoformat(),
            'coin': coin,
            'price': float(short_term_data.get('current_price', volume_data.get('current_price', 0))),
            'short_term': short_term_data,
            'volume_data': volume_data,
            'orderbook': orderbook_data,
            'indicators': indicators,
            'score': score,
            'signals': signals,
            'signal_type': signal_type
        }
    except Exception as e:
        return None

def analyze_short_term_volume(coin):
    """5분봉, 15분봉 기반 실시간 급등 감지"""
    try:
        df_5m = pyupbit.get_ohlcv(coin, interval="minute5", count=100)
        df_15m = pyupbit.get_ohlcv(coin, interval="minute15", count=100)
        
        if df_5m is None or df_15m is None or len(df_5m) < 20 or len(df_15m) < 20:
            return None
        
        # 5분봉 분석
        current_5m_volume = df_5m['volume'].iloc[-1]
        volume_5m_ma_10 = df_5m['volume'].rolling(10).mean().iloc[-1]
        volume_5m_ratio = current_5m_volume / volume_5m_ma_10 if volume_5m_ma_10 > 0 else 0
        
        recent_3_volume = df_5m['volume'].iloc[-3:].mean()
        prev_10_volume = df_5m['volume'].iloc[-13:-3].mean()
        volume_surge_ratio = recent_3_volume / prev_10_volume if prev_10_volume > 0 else 0
        
        price_change_5m = ((df_5m['close'].iloc[-1] - df_5m['close'].iloc[-4]) / df_5m['close'].iloc[-4]) * 100
        
        # 15분봉 분석
        current_15m_volume = df_15m['volume'].iloc[-1]
        volume_15m_ma_10 = df_15m['volume'].rolling(10).mean().iloc[-1]
        volume_15m_ratio = current_15m_volume / volume_15m_ma_10 if volume_15m_ma_10 > 0 else 0
        
        price_change_15m = ((df_15m['close'].iloc[-1] - df_15m['close'].iloc[-4]) / df_15m['close'].iloc[-4]) * 100
        
        # 연속 거래량 증가
        consecutive_increase = 0
        for i in range(1, min(5, len(df_5m))):
            if df_5m['volume'].iloc[-i] > df_5m['volume'].iloc[-i-1]:
                consecutive_increase += 1
            else:
                break
        
        # 체결강도 (양봉/음봉 비율)
        recent_candles = df_5m.iloc[-10:]
        bullish_count = sum(recent_candles['close'] > recent_candles['open'])
        bullish_ratio = bullish_count / 10
        
        return {
            'volume_5m_ratio': float(volume_5m_ratio),
            'volume_15m_ratio': float(volume_15m_ratio),
            'volume_surge_ratio': float(volume_surge_ratio),
            'price_change_5m': float(price_change_5m),
            'price_change_15m': float(price_change_15m),
            'consecutive_increase': int(consecutive_increase),
            'bullish_ratio': float(bullish_ratio),
            'current_price': float(df_5m['close'].iloc[-1])
        }
    except Exception as e:
        return None

def analyze_volume(coin):
    """거래량 분석 - 일봉 기반"""
    try:
        df = pyupbit.get_ohlcv(coin, interval="day", count=30)
        if df is None or len(df) < 20:
            return None
        
        current_volume = df['volume'].iloc[-1]
        volume_ma_20 = df['volume'].rolling(20).mean().iloc[-1]
        volume_ratio = current_volume / volume_ma_20
        
        volume_ma_7 = df['volume'].rolling(7).mean().iloc[-1]
        volume_ma_14 = df['volume'].rolling(14).mean().iloc[-1]
        accumulation_index = ((volume_ma_7 - volume_ma_14) / volume_ma_14) * 100
        
        price_7d_ago = df['close'].iloc[-8]
        current_price = df['close'].iloc[-1]
        price_change_7d = abs((current_price - price_7d_ago) / price_7d_ago) * 100
        
        price_change_1d = abs((df['close'].iloc[-1] - df['close'].iloc[-2]) / df['close'].iloc[-2]) * 100
        volume_change_1d = ((current_volume - df['volume'].iloc[-2]) / df['volume'].iloc[-2]) * 100
        
        divergence = volume_change_1d / price_change_1d if price_change_1d > 0 else 0
        
        return {
            'volume_ratio': float(volume_ratio),
            'accumulation_index': float(accumulation_index),
            'price_change_7d': float(price_change_7d),
            'divergence': float(divergence),
            'current_volume': float(current_volume),
            'current_price': float(current_price)
        }
    except Exception as e:
        return None

def analyze_orderbook(coin):
    """호가창 물량 변화 분석"""
    try:
        orderbook = pyupbit.get_orderbook(coin)
        if orderbook is None or not isinstance(orderbook, list) or len(orderbook) == 0:
            return None
        
        orderbook_data = orderbook[0]
        if 'orderbook_units' not in orderbook_data:
            return None
        
        units = orderbook_data['orderbook_units']
        total_bid_size = sum([item.get('bid_size', 0) for item in units])
        total_ask_size = sum([item.get('ask_size', 0) for item in units])
        bid_ask_ratio = total_bid_size / total_ask_size if total_ask_size > 0 else 0
        
        top_bid = units[0].get('bid_size', 0) if len(units) > 0 else 0
        top_ask = units[0].get('ask_size', 0) if len(units) > 0 else 0
        
        return {
            'total_bid': float(total_bid_size),
            'total_ask': float(total_ask_size),
            'bid_ask_ratio': float(bid_ask_ratio),
            'top_bid': float(top_bid),
            'top_ask': float(top_ask)
        }
    except Exception as e:
        return None

def calculate_indicators(coin):
    """기술적 지표 계산"""
    try:
        df = pyupbit.get_ohlcv(coin, interval="day", count=100)
        if df is None or len(df) < 50:
            return None
        
        rsi = ta.momentum.RSIIndicator(df['close'], window=14).rsi().iloc[-1]
        rsi_signal = "과매도" if rsi < 30 else "과매수" if rsi > 70 else "중립"
        
        macd = ta.trend.MACD(df['close'])
        macd_line = macd.macd().iloc[-1]
        signal_line = macd.macd_signal().iloc[-1]
        macd_hist = macd.macd_diff().iloc[-1]
        macd_signal = "골든크로스" if macd_line > signal_line and macd_hist > 0 else "데드크로스" if macd_line < signal_line and macd_hist < 0 else "중립"
        
        bollinger = ta.volatility.BollingerBands(df['close'])
        bb_high = bollinger.bollinger_hband().iloc[-1]
        bb_low = bollinger.bollinger_lband().iloc[-1]
        current_price = df['close'].iloc[-1]
        
        if current_price >= bb_high:
            bb_signal = "상단터치"
        elif current_price <= bb_low:
            bb_signal = "하단터치"
        else:
            bb_signal = "중립"
        
        ma5 = df['close'].rolling(5).mean().iloc[-1]
        ma20 = df['close'].rolling(20).mean().iloc[-1]
        ma_signal = "상향돌파" if ma5 > ma20 else "하향돌파"
        
        volume_avg = df['volume'].rolling(20).mean().iloc[-1]
        current_volume = df['volume'].iloc[-1]
        volume_percent = (current_volume / volume_avg) * 100
        volume_signal = "급증" if volume_percent > 150 else "정상"
        
        return {
            'rsi': float(rsi),
            'rsi_signal': rsi_signal,
            'macd_signal': macd_signal,
            'bb_signal': bb_signal,
            'ma_signal': ma_signal,
            'volume_percent': float(volume_percent),
            'volume_signal': volume_signal,
            'current_price': float(current_price)
        }
    except Exception as e:
        return None

# ============================================
# 신호 강도 판단
# ============================================

def calculate_signal_strength(volume_data, indicators, orderbook_data, short_term_data):
    """단기 + 중장기 지표 통합 분석 (최대 14개 지표)"""
    score = 0
    signals = []
    signal_type = "NORMAL"
    
    # 조기 감지 신호 (단기 시간봉)
    if short_term_data:
        if short_term_data['volume_5m_ratio'] >= 2.0:
            score += 2
            signals.append("🔥 5분봉 거래량 폭발")
            signal_type = "EARLY"
        elif short_term_data['volume_5m_ratio'] >= 1.5:
            score += 1
            signals.append("⚡ 5분봉 거래량 증가")
        
        if short_term_data['consecutive_increase'] >= 3:
            score += 2
            signals.append("🔥 연속 거래량 증가")
            signal_type = "EARLY"
        
        if short_term_data['price_change_5m'] > 5:
            score += 2
            signals.append("🚀 5분봉 급등 중")
            signal_type = "EARLY"
        elif short_term_data['price_change_5m'] > 3:
            score += 1
            signals.append("📈 5분봉 상승 중")
        
        if short_term_data['volume_15m_ratio'] >= 2.0:
            score += 1
            signals.append("✅ 15분봉 거래량 돌파")
        
        if short_term_data['bullish_ratio'] >= 0.7:
            score += 1
            signals.append("✅ 매수세 강함")
    
    # 일봉 거래량
    if volume_data:
        if volume_data['volume_ratio'] >= 2.0:
            score += 1
            signals.append("✅ 일봉 거래량 MA 돌파")
        
        if volume_data['accumulation_index'] > 20 and volume_data['price_change_7d'] < 5:
            score += 1
            signals.append("✅ 축적 패턴")
        
        if volume_data['divergence'] > 10:
            score += 1
            signals.append("✅ 고괴리")
    
    # 호가창
    if orderbook_data:
        if orderbook_data['bid_ask_ratio'] > 1.5:
            score += 1
            signals.append("✅ 매수벽 우세")
    
    # 기술적 지표
    if indicators:
        if indicators['rsi'] < 30:
            score += 1
            signals.append("✅ RSI 과매도")
        
        if indicators['macd_signal'] == "골든크로스":
            score += 1
            signals.append("✅ MACD 골든크로스")
        
        if indicators['bb_signal'] == "하단터치":
            score += 1
            signals.append("✅ 볼린저 하단")
        
        if indicators['ma_signal'] == "상향돌파":
            score += 1
            signals.append("✅ MA 상향돌파")
    
    return score, signals, signal_type

# ============================================
# 데이터 저장
# ============================================

def save_to_json_history(market_snapshot):
    """JSON 저장"""
    history_file = os.path.join(DATA_DIR, 'realtime_history.json')
    
    try:
        if os.path.exists(history_file):
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
        else:
            history = []
        
        history.append({
            'scan_time': get_kst_now().isoformat(),
            'data': market_snapshot
        })
        
        if len(history) > 100:
            history = history[-100:]
        
        with open(history_file, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        
        print(f"✅ JSON 저장: {len(market_snapshot)}개")
        return True
    except Exception as e:
        print(f"❌ JSON 저장 실패: {e}")
        return False

def save_to_excel_database(market_snapshot):
    """Excel 저장"""
    try:
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = "실시간모니터링"
            
            headers = ['수집시간', '코인', '신호타입', '점수', '현재가', '5분봉거래량', 
                      '가격변화5분', '연속증가', '일봉거래량', 'RSI', '판단']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        scan_time = get_kst_now().strftime('%Y-%m-%d %H:%M')
        for item in market_snapshot:
            short_term = item.get('short_term', {}) or {}
            volume_data = item.get('volume_data', {}) or {}
            indicators = item.get('indicators', {}) or {}
            
            row = [
                scan_time,
                item['coin'].replace('KRW-', ''),
                item['signal_type'],
                f"{item['score']}/14",
                item['price'],
                f"{short_term.get('volume_5m_ratio', 0):.2f}",
                f"{short_term.get('price_change_5m', 0):+.2f}%",
                f"{short_term.get('consecutive_increase', 0)}",
                f"{volume_data.get('volume_ratio', 0):.2f}",
                f"{indicators.get('rsi', 0):.1f}",
                "🔥조기감지" if item['signal_type'] == "EARLY" else "강력매수" if item['score'] >= 7 else "매수준비"
            ]
            ws.append(row)
        
        if ws.max_row > 1001:
            ws.delete_rows(2, ws.max_row - 1001)
        
        wb.save(EXCEL_FILE)
        print(f"✅ Excel 저장 완료")
        return True
    except Exception as e:
        print(f"❌ Excel 저장 실패: {e}")
        return False

def generate_realtime_report(market_snapshot):
    """실시간 모니터링 리포트"""
    report_date = get_kst_now().strftime('%Y%m%d_%H%M')
    report_path = os.path.join(ANALYSIS_DIR, f'realtime_report_{report_date}.md')
    
    signals = [(item, item['score'], item['signal_type']) for item in market_snapshot if item['score'] >= 4]
    signals.sort(key=lambda x: x[1], reverse=True)
    
    report = f"""# 실시간 모니터링 분석 리포트

생성시간: {get_kst_now().strftime('%Y-%m-%d %H:%M:%S')}

## 📊 스캔 요약

- 분석 코인 수: {len(market_snapshot)}개
- 신호 감지: {len(signals)}개
- 조기 감지: {sum([1 for _, _, t in signals if t == 'EARLY'])}개

## 🎯 주요 신호

"""
    
    for item, score, signal_type in signals[:20]:
        coin_name = item['coin'].replace('KRW-', '')
        short_term = item.get('short_term', {}) or {}
        
        report += f"""### {coin_name} (신호강도: {score}/14, {signal_type})

- 현재가: {item['price']:,.0f}원
- 5분봉 거래량: {short_term.get('volume_5m_ratio', 0):.2f}배
- 5분 가격변화: {short_term.get('price_change_5m', 0):+.2f}%
- 연속 증가: {short_term.get('consecutive_increase', 0)}회

"""
    
    report += f"""
## 📈 시장 통계

- 평균 5분봉 거래량 배수: {np.mean([item.get('short_term', {}).get('volume_5m_ratio', 0) for item in market_snapshot if item.get('short_term')]):.2f}
- 평균 5분 가격변화: {np.mean([item.get('short_term', {}).get('price_change_5m', 0) for item in market_snapshot if item.get('short_term')]):+.2f}%

---
*본 리포트는 자동 생성되었습니다.*
"""
    
    try:
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report)
        print(f"✅ 리포트 생성: {report_path}")
        return report_path, len(signals)
    except Exception as e:
        print(f"❌ 리포트 생성 실패: {e}")
        return None, 0

# ============================================
# Telegram & Git
# ============================================

def send_summary_notification(signals_count, early_count, report_path):
    """요약 알림"""
    if not BOT_TOKEN or not CHAT_ID:
        return
    
    try:
        message = f"""📊 실시간 모니터링 분석 완료

⏰ {get_kst_now().strftime('%Y-%m-%d %H:%M:%S')}
🎯 신호 감지: {signals_count}개
⚡ 조기 감지: {early_count}개
📈 리포트: {os.path.basename(report_path) if report_path else 'N/A'}

데이터는 Repository에 저장되었습니다."""

        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        data = {"chat_id": CHAT_ID, "text": message}
        requests.post(url, data=data, timeout=10)
    except Exception as e:
        print(f"알림 전송 실패: {e}")

def commit_and_push_data():
    """Git 커밋"""
    try:
        import subprocess
        
        subprocess.run(['git', 'config', '--global', 'user.email', 'github-actions[bot]@users.noreply.github.com'], check=True)
        subprocess.run(['git', 'config', '--global', 'user.name', 'GitHub Actions Bot'], check=True)
        
        subprocess.run(['git', 'add', DATA_DIR, ANALYSIS_DIR, EXCEL_FILE], check=True)
        
        commit_msg = f"Update realtime monitor - {get_kst_now().strftime('%Y-%m-%d %H:%M')}"
        result = subprocess.run(['git', 'commit', '-m', commit_msg], capture_output=True, text=True)
        
        if result.returncode == 0:
            subprocess.run(['git', 'push'], check=True)
            print("✅ Git 커밋 및 푸시 완료")
            return True
        else:
            print("ℹ️ 변경사항 없음")
            return False
            
    except Exception as e:
        print(f"⚠️ Git 작업 실패: {e}")
        return False

# ============================================
# 메인
# ============================================

def main():
    """메인"""
    print("""
    ╔══════════════════════════════════════╗
    ║   실시간 모니터링 분석 시스템       ║
    ║   Realtime Market Monitor           ║
    ╚══════════════════════════════════════╝
    """)
    
    try:
        market_snapshot = collect_market_data()
        
        if not market_snapshot:
            print("❌ 수집된 데이터 없음")
            return
        
        save_to_json_history(market_snapshot)
        save_to_excel_database(market_snapshot)
        
        report_path, signals_count = generate_realtime_report(market_snapshot)
        
        early_count = sum([1 for item in market_snapshot if item['signal_type'] == 'EARLY'])
        
        commit_and_push_data()
        
        send_summary_notification(signals_count, early_count, report_path)
        
        print(f"\n✅ 분석 완료: {len(market_snapshot)}개 코인, {signals_count}개 신호")
        
    except KeyboardInterrupt:
        print("\n🛑 프로그램 종료")
    except Exception as e:
        print(f"❌ 오류 발생: {e}")

if __name__ == "__main__":
    main()
